import csv
import io
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Literal
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import current_data_scope, current_tenant_id, current_user_id
from app.core.exceptions import AppError
from app.services.alert_rules import (
    generate_bad_review_alert_details_for_rows,
    generate_inventory_risk_alert_details_for_rows,
    generate_sales_drop_alert_details_for_rows,
)
from app.services.audit_service import insert_audit_log

router = APIRouter(prefix="/v1/imports", tags=["imports"])

ImportType = Literal["sales_daily", "sales_product", "inventory", "reviews"]

TemplateField = tuple[str, str, bool, str, str]

TEMPLATES: dict[str, list[TemplateField]] = {
    "sales_daily": [
        ("日期", "biz_date", True, "2026-06-10", "YYYY-MM-DD"),
        ("门店编码", "store_code", True, "SZ001", "系统内已存在门店编码"),
        ("门店名称", "store_name", False, "中山公园店", "辅助识别,不参与匹配"),
        ("渠道", "channel", True, "meituan", "meituan / eleme / offline 等"),
        ("实收营业额(元)", "revenue", True, "3580.50", "数字,不能为负"),
        ("订单数", "orders", True, "156", "整数,不能为负"),
        ("折扣金额(元)", "discount_amt", False, "320.00", "数字,默认 0"),
        ("退款金额(元)", "refund_amt", False, "45.00", "数字,默认 0"),
        ("备注", "note", False, "设备故障停业3h", "可选"),
    ],
    "sales_product": [
        ("日期", "biz_date", True, "2026-06-10", "YYYY-MM-DD"),
        ("门店编码", "store_code", True, "SZ001", "系统内已存在门店编码"),
        ("SKU编码", "sku", True, "MTC-001", "产品 SKU"),
        ("产品名称", "product_name", False, "杨枝甘露大杯", "辅助识别"),
        ("销售数量(杯)", "qty", True, "88", "整数,不能为负"),
        ("销售金额(元)", "revenue", True, "1496.00", "数字,不能为负"),
    ],
    "inventory": [
        ("日期", "biz_date", True, "2026-06-10", "YYYY-MM-DD"),
        ("门店编码", "store_code", True, "SZ001", "系统内已存在门店编码"),
        ("原料编码", "material_code", True, "RM-MANGO-01", "原料主数据编码"),
        ("原料名称", "material_name", True, "新鲜芒果", "原料名称"),
        ("单位", "unit", True, "kg", "kg / g / L / 个 等"),
        ("当日入库量", "inbound_qty", False, "5.00", "数字,默认 0"),
        ("当日消耗量", "usage_qty", False, "3.20", "数字,默认 0"),
        ("期末库存", "closing_stock", True, "12.50", "数字,不能为负"),
        ("安全库存线", "safety_stock", False, "8.00", "数字,默认 0"),
    ],
    "reviews": [
        ("评价时间", "created_at", True, "2026-06-10 14:23", "YYYY-MM-DD HH:mm"),
        ("门店编码", "store_code", True, "SZ001", "系统内已存在门店编码"),
        ("平台", "platform", True, "meituan", "meituan / eleme 等"),
        ("评分", "rating", True, "2.0", "0-5 数字"),
        ("评价内容", "content", False, "芒果不新鲜,有酸味", "可选"),
        ("是否已回复", "replied", False, "否", "是 / 否"),
    ],
}


class ImportConfirm(BaseModel):
    import_type: ImportType
    file_name: str = Field(min_length=1)
    total_rows: int = Field(default=0, ge=0)
    success_rows: int = Field(default=0, ge=0)
    overwrite_rows: int = Field(default=0, ge=0)
    mapping: dict[str, str] = Field(default_factory=dict)
    rows: list[dict[str, str]] = Field(default_factory=list)
    errors: list[dict] = Field(default_factory=list)


def template_for(import_type: str) -> list[TemplateField]:
    return TEMPLATES[import_type]


@router.get("/templates/{import_type}")
def download_template(import_type: ImportType) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"
    fields = template_for(import_type)
    sheet.append([column for column, _, _, _, _ in fields])
    sheet.append([example for _, _, _, example, _ in fields])
    note_sheet = workbook.create_sheet("字段说明")
    note_sheet.append(["模板列名", "字段标识", "是否必填", "格式说明"])
    for column, field, required, _, note in fields:
        note_sheet.append([column, field, "是" if required else "否", note])
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 28)
    for column_cells in note_sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        note_sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 28)
    output = io.BytesIO()
    workbook.save(output)
    filename = f"{import_type}_template.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/upload")
async def upload_import_file(
    request: Request,
    import_type: ImportType = Form(...),
    mapping_json: str | None = Form(default=None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    raw = await file.read()
    max_size = settings.max_upload_size_mb * 1024 * 1024
    if len(raw) > max_size:
        raise AppError(
            code="IMPORT_FILE_TOO_LARGE",
            message=f"文件超过 {settings.max_upload_size_mb}MB,请拆分后再导入",
            status_code=400,
        )

    rows, detected_headers = parse_uploaded_rows(file.filename, raw)
    fields = template_for(import_type)
    mapping = mapping_from_form(mapping_json, fields)
    expected_fields = [
        {"column": mapping.get(field) or column, "default_column": column, "field": field, "required": required, "note": note}
        for column, field, required, _, note in fields
    ]
    missing_required = [
        mapping.get(field) or column
        for column, field, required, _, _ in fields
        if required and (mapping.get(field) or column) not in detected_headers
    ]
    row_errors = validate_preview_rows(
        db=db,
        request=request,
        import_type=import_type,
        rows=rows,
        mapping=mapping,
        missing_required=missing_required,
    )
    valid_rows = 0 if missing_required else max(len(rows) - len({error["row"] for error in row_errors if error.get("row")}), 0)
    return {
        "import_type": import_type,
        "file_name": file.filename,
        "file_size": len(raw),
        "expected_fields": expected_fields,
        "detected_headers": detected_headers,
        "mapping": mapping,
        "missing_required": missing_required,
        "row_errors": row_errors,
        "valid": not missing_required and not row_errors,
        "preview_rows": rows[:5],
        "rows": rows,
        "total_rows": len(rows),
        "valid_rows": valid_rows,
        "total_preview_rows": min(len(rows), 5),
    }


@router.post("/confirm", status_code=201)
def confirm_import(payload: ImportConfirm, request: Request, db: Session = Depends(get_db)) -> dict:
    tenant_id = current_tenant_id(request)
    user_id = current_user_id(request)
    data_scope = current_data_scope(request)
    total_rows = payload.total_rows or len(payload.rows)
    success_rows = payload.success_rows
    overwrite_rows = payload.overwrite_rows
    alert_count = 0
    alert_details: list[dict] = []
    errors: list[dict] = []
    if payload.import_type == "sales_daily" and payload.rows:
        result = import_sales_daily_rows(db, tenant_id, user_id, data_scope, payload.mapping, payload.rows)
        total_rows = result["total_rows"]
        success_rows = result["success_rows"]
        overwrite_rows = result["overwrite_rows"]
        alert_count = result["alert_count"]
        alert_details = result.get("alert_details", [])
        errors.extend(result["errors"])
    elif payload.import_type == "sales_product" and payload.rows:
        result = import_sales_product_rows(db, tenant_id, user_id, data_scope, payload.mapping, payload.rows)
        total_rows = result["total_rows"]
        success_rows = result["success_rows"]
        overwrite_rows = result["overwrite_rows"]
        errors.extend(result["errors"])
    elif payload.import_type == "inventory" and payload.rows:
        result = import_inventory_rows(db, tenant_id, user_id, data_scope, payload.mapping, payload.rows)
        total_rows = result["total_rows"]
        success_rows = result["success_rows"]
        overwrite_rows = result["overwrite_rows"]
        alert_count = result["alert_count"]
        alert_details = result.get("alert_details", [])
        errors.extend(result["errors"])
    elif payload.import_type == "reviews" and payload.rows:
        result = import_review_rows(db, tenant_id, user_id, data_scope, payload.mapping, payload.rows)
        total_rows = result["total_rows"]
        success_rows = result["success_rows"]
        overwrite_rows = result["overwrite_rows"]
        alert_count = result["alert_count"]
        alert_details = result.get("alert_details", [])
        errors.extend(result["errors"])
    else:
        errors = list(payload.errors)
    status = "success"
    if errors and success_rows:
        status = "partial_success"
    elif errors:
        status = "failed"
    row = db.execute(
        text(
            """
            INSERT INTO import_jobs (
              tenant_id, user_id, import_type, file_url, status,
              total_rows, success_rows, overwrite_rows, error_details, metadata
            )
            VALUES (
              :tenant_id, :user_id, :import_type, :file_url, :status,
              :total_rows, :success_rows, :overwrite_rows,
              CAST(:error_details AS jsonb), CAST(:metadata AS jsonb)
            )
            RETURNING id, tenant_id, user_id, import_type, file_url, status,
                      total_rows, success_rows, overwrite_rows, error_details, metadata, created_at
            """
        ),
        {
            "tenant_id": tenant_id,
            "user_id": user_id,
            "import_type": payload.import_type,
            "file_url": payload.file_name,
            "status": status,
            "total_rows": total_rows,
            "success_rows": success_rows,
            "overwrite_rows": overwrite_rows,
            "error_details": json.dumps(errors, ensure_ascii=False),
            "metadata": json.dumps(
                {
                    "mapping": payload.mapping,
                    "generated_alerts": alert_count,
                    "generated_alerts_detail": alert_details,
                },
                ensure_ascii=False,
            ),
        },
    ).mappings().one()
    insert_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=user_id,
        action="IMPORT_CONFIRM",
        module="imports",
        object_type="import_job",
        object_id=row["id"],
        result=status,
        ip=request.client.host if request.client else None,
    )
    db.commit()
    return dict(row)


@router.get("/records")
def list_import_records(request: Request, db: Session = Depends(get_db)) -> list[dict]:
    tenant_id = current_tenant_id(request)
    rows = db.execute(
        text(
            """
            SELECT id, tenant_id, user_id, import_type, file_url, status,
                   total_rows, success_rows, overwrite_rows,
                   jsonb_array_length(COALESCE(error_details, '[]'::jsonb)) AS error_count,
                   metadata, created_at
            FROM import_jobs
            WHERE tenant_id = :tenant_id
            ORDER BY created_at DESC
            LIMIT 100
            """
        ),
        {"tenant_id": tenant_id},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.get("/records/{record_id}")
def get_import_record(record_id: UUID, request: Request, db: Session = Depends(get_db)) -> dict:
    tenant_id = current_tenant_id(request)
    row = db.execute(
        text(
            """
            SELECT ij.id, ij.tenant_id, ij.user_id, u.name AS user_name,
                   ij.import_type, ij.file_url, ij.status,
                   ij.total_rows, ij.success_rows, ij.overwrite_rows,
                   ij.error_details, ij.metadata, ij.created_at
            FROM import_jobs ij
            LEFT JOIN users u ON u.id = ij.user_id
            WHERE ij.id = :record_id AND ij.tenant_id = :tenant_id
            """
        ),
        {"tenant_id": tenant_id, "record_id": record_id},
    ).mappings().first()
    if not row:
        raise AppError(code="IMPORT_RECORD_NOT_FOUND", message="Import record not found", status_code=404)
    data = dict(row)
    errors = data.get("error_details") or []
    metadata = data.get("metadata") or {}
    data["error_count"] = len(errors)
    data["metadata"] = metadata
    data["mapping"] = metadata.get("mapping") or {}
    return data


@router.get("/records/{record_id}/errors")
def get_import_errors(record_id: UUID, request: Request, db: Session = Depends(get_db)) -> dict:
    tenant_id = current_tenant_id(request)
    row = db.execute(
        text(
            """
            SELECT id, error_details
            FROM import_jobs
            WHERE id = :record_id AND tenant_id = :tenant_id
            """
        ),
        {"tenant_id": tenant_id, "record_id": record_id},
    ).mappings().first()
    if not row:
        raise AppError(code="IMPORT_RECORD_NOT_FOUND", message="Import record not found", status_code=404)
    return {"id": row["id"], "errors": row["error_details"] or []}


@router.get("/records/{record_id}/errors/download")
def download_import_errors(record_id: UUID, request: Request, db: Session = Depends(get_db)) -> Response:
    tenant_id = current_tenant_id(request)
    row = db.execute(
        text(
            """
            SELECT id, import_type, file_url, error_details
            FROM import_jobs
            WHERE id = :record_id AND tenant_id = :tenant_id
            """
        ),
        {"tenant_id": tenant_id, "record_id": record_id},
    ).mappings().first()
    if not row:
        raise AppError(code="IMPORT_RECORD_NOT_FOUND", message="Import record not found", status_code=404)
    content = errors_to_csv(row["error_details"] or [])
    filename = f"{row['import_type']}_errors_{record_id}.csv"
    return Response(
        content=content.encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def mapping_from_form(mapping_json: str | None, fields: list[TemplateField]) -> dict[str, str]:
    default_mapping = {field: column for column, field, _, _, _ in fields}
    if not mapping_json:
        return default_mapping
    try:
        parsed = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise AppError(code="IMPORT_MAPPING_INVALID", message="字段映射不是有效 JSON", status_code=400) from exc
    if not isinstance(parsed, dict):
        raise AppError(code="IMPORT_MAPPING_INVALID", message="字段映射必须是对象", status_code=400)
    mapping = dict(default_mapping)
    for field, column in parsed.items():
        if field in mapping and isinstance(column, str) and column.strip():
            mapping[field] = column.strip()
    return mapping


def errors_to_csv(errors: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["row", "code", "message", "detail"], lineterminator="\n")
    writer.writeheader()
    for error in errors:
        writer.writerow(
            {
                "row": error.get("row", ""),
                "code": error.get("code", ""),
                "message": error.get("message", ""),
                "detail": json.dumps(error.get("detail") or {}, ensure_ascii=False),
            }
        )
    return output.getvalue()


def parse_uploaded_rows(file_name: str, raw: bytes) -> tuple[list[dict[str, str]], list[str]]:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        text_body = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_body))
        headers = [normalize_cell(header) for header in (reader.fieldnames or []) if normalize_cell(header)]
        rows = [{normalize_cell(key): normalize_cell(value) for key, value in row.items() if normalize_cell(key)} for row in reader]
        return rows, headers
    if lower_name.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        headers = [normalize_cell(value) for value in (header_row or [])]
        rows: list[dict[str, str]] = []
        for values in iterator:
            if not values or all(normalize_cell(value) == "" for value in values):
                continue
            row = {
                header: normalize_cell(value)
                for header, value in zip(headers, values)
                if header
            }
            rows.append(row)
        return rows, [header for header in headers if header]
    raise AppError(
        code="IMPORT_TEMPLATE_INVALID",
        message="仅支持 CSV 或 XLSX 文件",
        status_code=422,
    )


def validate_preview_rows(
    db: Session,
    request: Request,
    import_type: str,
    rows: list[dict[str, str]],
    mapping: dict[str, str],
    missing_required: list[str],
) -> list[dict]:
    if missing_required:
        return [
            {
                "row": 1,
                "field": "header",
                "code": "IMPORT_TEMPLATE_INVALID",
                "message": f"缺少必填列:{', '.join(missing_required)}",
            }
        ]
    if not rows:
        return [{"row": 2, "field": "rows", "code": "IMPORT_EMPTY_FILE", "message": "文件没有可导入的数据行"}]

    tenant_id = current_tenant_id(request)
    user_id = current_user_id(request)
    data_scope = current_data_scope(request)
    errors: list[dict] = []
    for index, row in enumerate(rows, start=2):
        try:
            validate_row_by_type(db, tenant_id, user_id, data_scope, import_type, mapping, row)
        except AppError as exc:
            errors.append({"row": index, "code": exc.code, "message": exc.message, "detail": exc.detail})
    return errors[:200]


def validate_row_by_type(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    import_type: str,
    mapping: dict[str, str],
    row: dict[str, str],
) -> None:
    if import_type == "sales_daily":
        parse_sales_daily_row(db, tenant_id, user_id, data_scope, mapping, row)
        return
    if import_type == "sales_product":
        ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
        parse_date(required_value(row, mapping, "biz_date"), "biz_date")
        ensure_product_exists(db, tenant_id, required_value(row, mapping, "sku"))
        integer_value(row, mapping, "qty")
        decimal_value(row, mapping, "revenue")
        return
    if import_type == "inventory":
        ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
        parse_date(required_value(row, mapping, "biz_date"), "biz_date")
        required_value(row, mapping, "material_code")
        required_value(row, mapping, "material_name")
        required_value(row, mapping, "unit")
        decimal_value(row, mapping, "inbound_qty", default="0")
        decimal_value(row, mapping, "usage_qty", default="0")
        decimal_value(row, mapping, "closing_stock")
        decimal_value(row, mapping, "safety_stock", default="0")
        return
    if import_type == "reviews":
        ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
        parse_datetime(required_value(row, mapping, "created_at"), "created_at")
        required_value(row, mapping, "platform")
        rating = decimal_value(row, mapping, "rating")
        if rating < 1 or rating > 5:
            raise AppError(code="IMPORT_RATING_INVALID", message="评分必须在 1 到 5 之间", detail={"field": "rating"})


def import_sales_daily_rows(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    rows: list[dict[str, str]],
) -> dict:
    aggregates: dict[tuple[str, str], dict] = {}
    errors: list[dict] = []
    for index, raw_row in enumerate(rows, start=2):
        try:
            parsed = parse_sales_daily_row(db, tenant_id, user_id, data_scope, mapping, raw_row)
        except AppError as exc:
            errors.append({"row": index, "code": exc.code, "message": exc.message, "detail": exc.detail})
            continue
        key = (parsed["store_id"], parsed["biz_date"])
        if key not in aggregates:
            aggregates[key] = {
                "store_id": parsed["store_id"],
                "biz_date": parsed["biz_date"],
                "revenue": Decimal("0"),
                "orders": 0,
                "discount_amt": Decimal("0"),
                "refund_amt": Decimal("0"),
                "channel_json": {},
            }
        item = aggregates[key]
        item["revenue"] += parsed["revenue"]
        item["orders"] += parsed["orders"]
        item["discount_amt"] += parsed["discount_amt"]
        item["refund_amt"] += parsed["refund_amt"]
        item["channel_json"][parsed["channel"]] = str(
            Decimal(item["channel_json"].get(parsed["channel"], "0")) + parsed["revenue"]
        )

    overwrite_rows = 0
    for item in aggregates.values():
        existed = db.execute(
            text(
                """
                SELECT 1
                FROM sales_daily
                WHERE tenant_id = :tenant_id AND store_id = :store_id AND biz_date = :biz_date
                """
            ),
            {"tenant_id": tenant_id, "store_id": item["store_id"], "biz_date": item["biz_date"]},
        ).scalar()
        if existed:
            overwrite_rows += 1
        avg_order = item["revenue"] / item["orders"] if item["orders"] else Decimal("0")
        db.execute(
            text(
                """
                INSERT INTO sales_daily (
                  tenant_id, store_id, biz_date, revenue, orders, avg_order,
                  discount_amt, refund_amt, channel_json
                )
                VALUES (
                  :tenant_id, :store_id, :biz_date, :revenue, :orders, :avg_order,
                  :discount_amt, :refund_amt, CAST(:channel_json AS jsonb)
                )
                ON CONFLICT (tenant_id, store_id, biz_date) DO UPDATE
                SET revenue = EXCLUDED.revenue,
                    orders = EXCLUDED.orders,
                    avg_order = EXCLUDED.avg_order,
                    discount_amt = EXCLUDED.discount_amt,
                    refund_amt = EXCLUDED.refund_amt,
                    channel_json = EXCLUDED.channel_json
                """
            ),
            {
                "tenant_id": tenant_id,
                "store_id": item["store_id"],
                "biz_date": item["biz_date"],
                "revenue": item["revenue"],
                "orders": item["orders"],
                "avg_order": avg_order,
                "discount_amt": item["discount_amt"],
                "refund_amt": item["refund_amt"],
                "channel_json": json.dumps(item["channel_json"]),
            },
        )
    alert_details = generate_sales_drop_alert_details_for_rows(db, tenant_id, user_id, list(aggregates.values()))
    return {
        "total_rows": len(rows),
        "success_rows": len(rows) - len(errors),
        "overwrite_rows": overwrite_rows,
        "alert_count": len(alert_details),
        "alert_details": alert_details,
        "errors": errors,
    }


def import_sales_product_rows(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    rows: list[dict[str, str]],
) -> dict:
    errors: list[dict] = []
    parsed_rows: list[dict] = []
    overwrite_rows = 0
    for index, raw_row in enumerate(rows, start=2):
        try:
            parsed = parse_sales_product_row(db, tenant_id, user_id, data_scope, mapping, raw_row)
        except AppError as exc:
            errors.append({"row": index, "code": exc.code, "message": exc.message, "detail": exc.detail})
            continue
        parsed_rows.append(parsed)

    for item in parsed_rows:
        existed = db.execute(
            text(
                """
                SELECT 1
                FROM sales_product_daily
                WHERE tenant_id = :tenant_id
                  AND store_id = :store_id
                  AND product_id = :product_id
                  AND biz_date = :biz_date
                """
            ),
            {
                "tenant_id": tenant_id,
                "store_id": item["store_id"],
                "product_id": item["product_id"],
                "biz_date": item["biz_date"],
            },
        ).scalar()
        if existed:
            overwrite_rows += 1
        db.execute(
            text(
                """
                INSERT INTO sales_product_daily (
                  tenant_id, store_id, product_id, biz_date, qty, revenue, margin
                )
                VALUES (
                  :tenant_id, :store_id, :product_id, :biz_date, :qty, :revenue, :margin
                )
                ON CONFLICT (tenant_id, store_id, product_id, biz_date) DO UPDATE
                SET qty = EXCLUDED.qty,
                    revenue = EXCLUDED.revenue,
                    margin = EXCLUDED.margin
                """
            ),
            {"tenant_id": tenant_id, **item},
        )
    return {
        "total_rows": len(rows),
        "success_rows": len(parsed_rows),
        "overwrite_rows": overwrite_rows,
        "alert_count": 0,
        "errors": errors,
    }


def import_inventory_rows(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    rows: list[dict[str, str]],
) -> dict:
    errors: list[dict] = []
    parsed_rows: list[dict] = []
    overwrite_rows = 0
    for index, raw_row in enumerate(rows, start=2):
        try:
            parsed = parse_inventory_row(db, tenant_id, user_id, data_scope, mapping, raw_row)
        except AppError as exc:
            errors.append({"row": index, "code": exc.code, "message": exc.message, "detail": exc.detail})
            continue
        parsed_rows.append(parsed)

    for item in parsed_rows:
        existed = db.execute(
            text(
                """
                SELECT 1
                FROM inventory_snapshots
                WHERE tenant_id = :tenant_id
                  AND store_id = :store_id
                  AND material_id = :material_id
                  AND biz_date = :biz_date
                """
            ),
            {
                "tenant_id": tenant_id,
                "store_id": item["store_id"],
                "material_id": item["material_id"],
                "biz_date": item["biz_date"],
            },
        ).scalar()
        if existed:
            overwrite_rows += 1
        db.execute(
            text(
                """
                INSERT INTO inventory_snapshots (
                  tenant_id, store_id, material_id, biz_date,
                  inbound_qty, usage_qty, closing_stock
                )
                VALUES (
                  :tenant_id, :store_id, :material_id, :biz_date,
                  :inbound_qty, :usage_qty, :closing_stock
                )
                ON CONFLICT (tenant_id, store_id, material_id, biz_date) DO UPDATE
                SET inbound_qty = EXCLUDED.inbound_qty,
                    usage_qty = EXCLUDED.usage_qty,
                    closing_stock = EXCLUDED.closing_stock
                """
            ),
            {"tenant_id": tenant_id, **item},
        )
    alert_details = generate_inventory_risk_alert_details_for_rows(db, tenant_id, user_id, parsed_rows)
    return {
        "total_rows": len(rows),
        "success_rows": len(parsed_rows),
        "overwrite_rows": overwrite_rows,
        "alert_count": len(alert_details),
        "alert_details": alert_details,
        "errors": errors,
    }


def import_review_rows(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    rows: list[dict[str, str]],
) -> dict:
    errors: list[dict] = []
    parsed_rows: list[dict] = []
    for index, raw_row in enumerate(rows, start=2):
        try:
            parsed = parse_review_row(db, tenant_id, user_id, data_scope, mapping, raw_row)
        except AppError as exc:
            errors.append({"row": index, "code": exc.code, "message": exc.message, "detail": exc.detail})
            continue
        parsed_rows.append(parsed)

    for item in parsed_rows:
        db.execute(
            text(
                """
                INSERT INTO reviews (
                  tenant_id, store_id, platform, rating, content,
                  replied, sentiment, category, created_at
                )
                VALUES (
                  :tenant_id, :store_id, :platform, :rating, :content,
                  :replied, :sentiment, :category, :created_at
                )
                """
            ),
            {"tenant_id": tenant_id, **item},
        )
    alert_details = generate_bad_review_alert_details_for_rows(db, tenant_id, user_id, parsed_rows)
    return {
        "total_rows": len(rows),
        "success_rows": len(parsed_rows),
        "overwrite_rows": 0,
        "alert_count": len(alert_details),
        "alert_details": alert_details,
        "errors": errors,
    }


def parse_sales_daily_row(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    row: dict[str, str],
) -> dict:
    store_code = required_value(row, mapping, "store_code")
    store_id = ensure_store_exists(db, tenant_id, user_id, data_scope, store_code)
    biz_date = parse_date(required_value(row, mapping, "biz_date"), "biz_date")
    orders = integer_value(row, mapping, "orders")
    return {
        "store_id": store_id,
        "biz_date": biz_date,
        "channel": required_value(row, mapping, "channel"),
        "revenue": decimal_value(row, mapping, "revenue"),
        "orders": orders,
        "discount_amt": decimal_value(row, mapping, "discount_amt", default="0"),
        "refund_amt": decimal_value(row, mapping, "refund_amt", default="0"),
    }


def parse_sales_product_row(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    row: dict[str, str],
) -> dict:
    store_id = ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
    product = ensure_product_exists(db, tenant_id, required_value(row, mapping, "sku"))
    qty = integer_value(row, mapping, "qty")
    revenue = decimal_value(row, mapping, "revenue")
    cost = product.get("cost")
    margin = None
    if cost is not None:
        margin = revenue - Decimal(str(cost)) * qty
    return {
        "store_id": store_id,
        "product_id": product["id"],
        "biz_date": parse_date(required_value(row, mapping, "biz_date"), "biz_date"),
        "qty": qty,
        "revenue": revenue,
        "margin": margin,
    }


def parse_inventory_row(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    row: dict[str, str],
) -> dict:
    store_id = ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
    material = ensure_material(
        db,
        tenant_id,
        material_code=required_value(row, mapping, "material_code"),
        name=required_value(row, mapping, "material_name"),
        unit=required_value(row, mapping, "unit"),
        safety_stock=optional_decimal_value(row, mapping, "safety_stock"),
    )
    return {
        "store_id": store_id,
        "material_id": material["id"],
        "biz_date": parse_date(required_value(row, mapping, "biz_date"), "biz_date"),
        "inbound_qty": decimal_value(row, mapping, "inbound_qty", default="0"),
        "usage_qty": optional_decimal_value(row, mapping, "usage_qty"),
        "closing_stock": decimal_value(row, mapping, "closing_stock"),
    }


def parse_review_row(
    db: Session,
    tenant_id: str,
    user_id: str,
    data_scope: str,
    mapping: dict[str, str],
    row: dict[str, str],
) -> dict:
    store_id = ensure_store_exists(db, tenant_id, user_id, data_scope, required_value(row, mapping, "store_code"))
    rating = decimal_value(row, mapping, "rating")
    if rating < 1 or rating > 5:
        raise AppError(code="IMPORT_RATING_INVALID", message="评分必须在 1 到 5 之间", detail={"field": "rating"})
    content = value_for(row, mapping, "content", "")
    sentiment, category = classify_review(rating, content)
    return {
        "store_id": store_id,
        "platform": required_value(row, mapping, "platform"),
        "rating": rating,
        "content": content,
        "replied": boolean_value(row, mapping, "replied", default=False),
        "sentiment": sentiment,
        "category": category,
        "created_at": parse_datetime(required_value(row, mapping, "created_at"), "created_at"),
    }


def ensure_store_exists(db: Session, tenant_id: str, user_id: str, data_scope: str, store_code: str):
    scope_sql = "1=1"
    params = {"tenant_id": tenant_id, "code": store_code}
    if data_scope != "all":
        scope_sql = """
            (
              manager_user_id = :user_id
              OR franchisee_user_id = :user_id
              OR EXISTS (
                SELECT 1
                FROM user_store_scopes uss
                WHERE uss.tenant_id = stores.tenant_id
                  AND uss.user_id = :user_id
                  AND uss.store_id = stores.id
              )
            )
        """
        params["user_id"] = user_id
    store_id = db.execute(
        text(
            f"""
            SELECT id
            FROM stores
            WHERE tenant_id = :tenant_id
              AND code = :code
              AND {scope_sql}
            """
        ),
        params,
    ).scalar()
    if not store_id:
        raise AppError(
            code="STORE_NOT_FOUND",
            message=f"门店编码不存在或无权限访问:{store_code}",
            detail={"store_code": store_code},
        )
    return store_id


def ensure_product_exists(db: Session, tenant_id: str, sku: str) -> dict:
    product = db.execute(
        text(
            """
            SELECT id, sku, name, cost
            FROM products
            WHERE tenant_id = :tenant_id AND sku = :sku AND status = 'active'
            """
        ),
        {"tenant_id": tenant_id, "sku": sku},
    ).mappings().first()
    if not product:
        raise AppError(code="PRODUCT_NOT_FOUND", message=f"SKU 不存在或未启用:{sku}", detail={"sku": sku})
    return dict(product)


def ensure_material(
    db: Session,
    tenant_id: str,
    *,
    material_code: str,
    name: str,
    unit: str,
    safety_stock: Decimal | None,
) -> dict:
    row = db.execute(
        text(
            """
            INSERT INTO materials (tenant_id, material_code, name, unit, safety_stock)
            VALUES (:tenant_id, :material_code, :name, :unit, :safety_stock)
            ON CONFLICT (tenant_id, material_code) DO UPDATE
            SET name = EXCLUDED.name,
                unit = EXCLUDED.unit,
                safety_stock = COALESCE(EXCLUDED.safety_stock, materials.safety_stock)
            RETURNING id, tenant_id, material_code, name, unit, safety_stock, status
            """
        ),
        {
            "tenant_id": tenant_id,
            "material_code": material_code,
            "name": name,
            "unit": unit,
            "safety_stock": safety_stock,
        },
    ).mappings().one()
    return dict(row)


def value_for(row: dict[str, str], mapping: dict[str, str], field: str, default: str | None = None) -> str:
    column = mapping.get(field)
    value = row.get(column, default) if column else default
    return normalize_cell(value)


def required_value(row: dict[str, str], mapping: dict[str, str], field: str) -> str:
    value = value_for(row, mapping, field)
    if not value:
        raise AppError(code="IMPORT_FIELD_REQUIRED", message=f"缺少必填字段:{field}", detail={"field": field})
    return value


def decimal_value(row: dict[str, str], mapping: dict[str, str], field: str, default: str = "0") -> Decimal:
    raw_value = value_for(row, mapping, field, default)
    try:
        value = Decimal(raw_value or default)
    except InvalidOperation as exc:
        raise AppError(code="IMPORT_NUMBER_INVALID", message=f"数字格式不正确:{field}", detail={"field": field}) from exc
    if value < 0:
        raise AppError(code="IMPORT_NUMBER_NEGATIVE", message=f"数字不能小于 0:{field}", detail={"field": field})
    return value


def optional_decimal_value(row: dict[str, str], mapping: dict[str, str], field: str) -> Decimal | None:
    raw_value = value_for(row, mapping, field, None)
    if not raw_value:
        return None
    try:
        value = Decimal(raw_value)
    except InvalidOperation as exc:
        raise AppError(code="IMPORT_NUMBER_INVALID", message=f"数字格式不正确:{field}", detail={"field": field}) from exc
    if value < 0:
        raise AppError(code="IMPORT_NUMBER_NEGATIVE", message=f"数字不能小于 0:{field}", detail={"field": field})
    return value


def integer_value(row: dict[str, str], mapping: dict[str, str], field: str) -> int:
    raw_value = required_value(row, mapping, field)
    try:
        value = int(raw_value)
    except ValueError as exc:
        raise AppError(code="IMPORT_INTEGER_INVALID", message=f"整数格式不正确:{field}", detail={"field": field}) from exc
    if value < 0:
        raise AppError(code="IMPORT_INTEGER_NEGATIVE", message=f"整数不能小于 0:{field}", detail={"field": field})
    return value


def parse_date(value: str, field: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise AppError(code="IMPORT_DATE_INVALID", message="日期格式应为 YYYY-MM-DD", detail={"field": field}) from exc


def parse_datetime(value: str, field: str):
    for pattern in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    raise AppError(code="IMPORT_DATETIME_INVALID", message="时间格式应为 YYYY-MM-DD HH:mm", detail={"field": field})


def boolean_value(row: dict[str, str], mapping: dict[str, str], field: str, default: bool = False) -> bool:
    raw_value = value_for(row, mapping, field)
    if not raw_value:
        return default
    normalized = raw_value.lower()
    if normalized in {"是", "true", "1", "yes", "y", "已回复"}:
        return True
    if normalized in {"否", "false", "0", "no", "n", "未回复"}:
        return False
    raise AppError(code="IMPORT_BOOLEAN_INVALID", message=f"布尔字段只能填写是/否:{field}", detail={"field": field})


def classify_review(rating: Decimal, content: str) -> tuple[str, str | None]:
    text = content or ""
    if rating <= Decimal("2.0"):
        sentiment = "negative"
    elif rating >= Decimal("4.0"):
        sentiment = "positive"
    else:
        sentiment = "neutral"

    keyword_categories = (
        ("food_safety", ("食安", "食品安全", "异物", "呕吐", "腹泻", "变质", "发酸", "不新鲜")),
        ("taste", ("口味", "太甜", "太淡", "难喝", "不好喝", "酸", "苦")),
        ("service", ("服务", "态度", "客服", "店员")),
        ("packaging", ("包装", "撒了", "漏了", "破损")),
        ("speed", ("慢", "出餐", "配送", "等太久")),
    )
    for category, keywords in keyword_categories:
        if any(keyword in text for keyword in keywords):
            return sentiment, category
    return sentiment, "other" if sentiment == "negative" else None


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
